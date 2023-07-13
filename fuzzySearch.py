import openpyxl                                                     #this and below interacts w/ excel
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import sys
from sys import platform                                            #checks if on iOS or Windows
from fuzzywuzzy import fuzz                                         #helps search find potential matches
import re                                                           #regex, helps clean the names
import pickle                                                       #helps save/export data from the search

"""
###To learn about openpyxl library watch
https://www.youtube.com/watch?v=7YS6YDQKFh0

or watch this series
https://www.youtube.com/watch?v=6QdeR15myIY&list=PLCC34OHNcOtrMWIf_MXWrCajQwp3lL27T

###To learn about fuzzywuzzy go here:
https://www.datacamp.com/tutorial/fuzzy-string-python

###To learn how to save/load files from the pickle module go here:
https://stackoverflow.com/questions/20716812/saving-and-loading-multiple-objects-in-pickle-file

"""

"""introduction"""
print("\nHello! Let's check for duplicates.\n")                     #introduction at Terminal/CMD


"""global variables"""
file_path = ''                                                      #var to store file path
file_name = 'Accounts_and_Contacts'                                 #var to store file name
if platform == "win32":                                             #checks if OS is Windows
    file = 'D:\IT\Tasks\Python\Duplicate_Accts\\' + file_name + '.xlsx'
if platform == "darwin":                                            #checks if OS is iOS
    file = '/Users/Oscar/Desktop/DPA/Tasks/Python/Duplicate_Accts/' + file_name + '.xlsx'

print("Loading Excel workbook... \n")
wb = load_workbook(file)                                            #Opening excel workbook (wb) from path
ws = wb.active                                                      #creating an active worksheet (ws) from the workbook
threshhold = 93                                                     #number used to determine accuracy of Fuzzy search

###########################################################################################################################
"""A class of Accounts to save the full name, acct index in the excel sheet, and its matches"""
class Account():
    def __init__(self, first_name = "", middle_name = "", last_name = "", matches = [], index = 0):
        self.index = index                                          #index in excel
        self.matches = matches                                      #list of matches
        self.first_name = first_name                                #first name
        self.middle_name = middle_name                              #middle name
        self.last_name = last_name                                  #last name
        self.full_name = ""
        self.excel_name = ""

    def __index__():
        return self.index

    def fullname():
        fullname = self.first_name + " " + self.middle_name + " " + self.last_name
        self.full_name = fullname
        return fullname
    
    def match(self, match):                                         #stores a NEW match
        self.matches += [match]

    def total_matches(self):                                        #shows # of matches
        return len(self.matches)

    def info(self):
        vars(self)


###########################################################################################################################
"""Creates a file to access data after its creation."""
def to_pickle(data, filename = ""):
    with open(filename, "ab") as pickle_out:                        #creating output file in binary "ab"
        pickle.dump(data, pickle_out)                               #creating source/destination file
    pickle_out.close()


###########################################################################################################################
"""Helps read data that was pickled"""
def unpickle(filename = ""):
    data = open(filename, "rb")                                     #opening file to load data
    return pickle.load(data)
    data.close()
    print("File data has been unpickled...")


###########################################################################################################################
"""Will show the percentage done to nearest tenth, e.g. '12.4% complete' """
def percent_complete(i = 1, length = 1):
    print("   Fuzzy search is {:0.01f}% complete...".format(i*100/length), end = "\r")


###########################################################################################################################
def get_column(column = 'A'):
    col = ws[column + '2' :column + str(len(ws[column]))]
    cells = []                                                      #list to store all cells in the column

    for tuple in col:                                               #NOTE: cells are saved as tuples
        for cell in tuple:                                          #grabbing cell from tuple
            cells += [cell.value]                                   #clearing punctuations from cell.value
    return cells


###########################################################################################################################
"""This will help clear all punctuations from a string"""
def clear_punctuation(string = ""):
    if string is None:                                              #turns a None type into empty space string, " "
        return " "
    if string == "None":
        return " "
    else:                                                           #use Regex to capture everything but whitespaces
        return re.sub(r'[^\w]', '', string)

###########################################################################################################################
"""This will delete " None " from the names"""
def clear_none(string = ""):
    if "None" in string:
        return re.sub(r'None', '', string)
    return string

###########################################################################################################################
""" By default returns column A of excel file """
def get_selection(list = [], start = 0, end = 0):
    return list[start-2:end+1]                                      #we subtract 2 to "convert" from excel index


###########################################################################################################################
""" This method allows users to determin the value for the accuracy of the Fuzzy search."""
def get_threshold():
    getting_threshold = True

    while getting_threshold:
        try:
            value = int(input("\nWhat threshold do you want? 93 is recommended. Choose between 1 through 100:  "))
            if value > 100 or value < 0:
                print("   That's not an option, please choose between 1 and 100.")
            else:
                getting_threshold = False
        except ValueError:
            print("   Oops! That wasn't a number. Please enter a number.")
    return value


###########################################################################################################################
""" This method will add columns to the excel file"""
def insert_column(matches = [], column = 'A', start = 1, end = 1, header = ""):
    bold = Font(bold=True)                                          #pre-setting font to bold
    ws[str(column)+'1'] = header                                    #saving header cell
    ws[str(column)+'1'].font = bold                                 #making header bold

    #adding values to column
    index = 0
    for row in range(start,end+1):                                  #update range from 'start' to 'end', e.g.
        names = []
        for match in matches[index]:
            names += [clear_none(match)]
        index += 1
        ws.cell(row, ord(column)-64, ", ".join(names))
        
        
###########################################################################################################################
"""This method uses the fuzzywuzzy module to search for duplicate names"""
def fuzzy_check(all_accounts = [], all_names = [], f = [], m = [], l = [], start = 0, length = 1):
    all_accts = all_accounts                                        #list to store account's
    all_fullnames = []                                              #list to store full name details
    selected_fullnames = []                                         #list to store selected full names
    matches = [[] for _ in range(length)]                           #list to store matches of selected names
    i = 0                                                           #pointer index for matches list
    threshold = get_threshold()                                     #allow user to set the Fuzzy search accuracy

    print("   Starting fuzzy search...")
    #filling out seperate name info
    selected_fullnames = [ str(f[j]) + " " + str(m[j]) + " " + str(l[j]) for j in range(length) ]   #adding ~selected~ full names to list
    all_fullnames = [ account.excel_name for account in all_accounts ]                              #adding ~excel~ fullnames to list
    
    #searching for duplicates w/ Fuzzy Search
    
    for name in selected_fullnames:                                 #name to compare
        for comparison in all_fullnames:                            #chosen comparison for 'name'
            if name.index == comparison.index:                      #skip checking names against themselves by looking to their index
                continue
            sort_ratio = fuzz.token_sort_ratio(name, comparison)    #comparing name against comparison name w/ Fuzzy Search       
            if sort_ratio > threshhold :                                     #93 ratio determined after testing, certifies names are matches
                matches[i] += [comparison]                        #adding matches

        i += 1
        percent_complete(i, length)                              #shows percent complete with duplicate search
        
    return [selected_fullnames, matches]


###########################################################################################################################
"""Beginning the ~process~ of getting names and finding duplicates within the data."""
def start_program(start = 1, end = 1):
    print("\nRetreiving account information...")
    
    try:                                                            #We'll try to unpickle all names & acct info.
        all_accounts = unpickle("all_accounts")
        all_names = unpickle("all_names")
        all_first_names = all_names[0]
        all_middle_names = all_names[1]
        all_last_names = all_names[2]

        print("   Unpickling all names...")
    except FileNotFoundError:                                       #Otherwise, if files aren't found we create the name and acct info
        all_first_names = get_column('A')                           #gather information from excel
        all_middle_names = get_column('B')                       
        all_last_names = get_column('C')
        all_names = [all_first_names, all_middle_names, all_last_names]
        all_accounts = []                                           #list storing account information using the Account() class

        for i in range(len(all_first_names)):                       #From here down, we fill out the Account() class information
            account = Account()
            name = str(all_first_names[i]) + " " + str(all_middle_names[i]) + " " + str(all_last_names[i])      #saving name
            account.index = i+2                                                             #saving the ~excel~ index
            account.excel_name = name                                                       #saving the ~excel~ name
            account.first_name = clear_punctuation(all_first_names[i])                      #saving name data w/o punctuations or whitespaces
            account.middle_name = clear_punctuation(all_middle_names[i])
            account.last_name = clear_punctuation(all_last_names[i])
            account.full_name = " ".join([account.first_name, account.middle_name, account.last_name])
            all_accounts += [ account ]                                                     #saving account info to list
        to_pickle( [all_first_names, all_middle_names, all_last_names], "all_names")        #pickling all name info
        to_pickle( all_accounts, "all_accounts")                                            #pickling list of account info

    """here we retreive user's selection of names"""
    selected_first_names = get_selection(all_first_names, start, end)                       #getting the first names
    print("   Done retreiving names...")
    selected_middle_names = get_selection(all_middle_names, start, end)                     #getting the middle names
    print("   Done retreiving middle names...")
    selected_last_names = get_selection(all_last_names, start, end)                         #getting the last names
    print("   Done retreiving last names...")

    #here we start Fuzzy Search
    selected_data = fuzzy_check(all_accounts, all_names, selected_first_names, selected_middle_names, selected_last_names, start, end-start+1) #checking for duplicates

    #saving fullnames and matches to files named "fullnames" and "matches"
    """print("\n   Pickling items...")
    selection = str(start) + "_to_" + str(end)
    to_pickle(selected_data[0], "fullnames_" + selection)                 #pickling list of names
    to_pickle(selected_data[1], "matches_" + selection)                   #pickling list of selection_matches"""

    print("\n\nUpdating Excel document...")
    insert_column(selected_data[1], column = "G", start = start, end = end, header = "Matches")            #adding matches to excel in column G i.e., column 7

    #Closing and saving the Excel WorkBook
    #wb.save(file_name + "_updated" + "_rows_" + str(start) + "_to_" + str(end) + ".xlsx")   #saving updated workbook
    wb.save(file_name + ".xlsx")
    wb.close()                                                                              #closing the excel file


###########################################################################################################################
"""Provides an option to rerun the program."""
def start_again(start_index = 1, last_index = 1):
    start_again = input("\nDo you want to go through more names? (y/n) ")
    while start_again not in ["y", "n"]:
        print("\n   Please type y or n.")
        start_again = input("\nDo you want to go through more names? (y/n) ")
    if start_again == "y":
        print("\n   Last time you looked through cells", start_index, "to", last_index, "\n")
        _main()
    else:
        print("\nProgram has ended. Good-bye!")










#--------------------------------------DON'T DELETE ANY OF THE CODE BELOW-------------------------------------------------------
def _main():
    processing_names = True                                         #program hasn't run once
    making_index_choice = True                                      #used to check if user wants ALL indices or selection
    getting_index = True                                            #used to check if user has entered indices to terminal
    got_start_index = False                                         #used to check if user entered start excel index                      
    got_last_index = False                                          #used to check if user entered last excel index

    while getting_index:                                            #prompting user to enter excel indices to check for names
        try:
            if not got_start_index:                                 #first excel cell to capture
                start_index = int(input("Enter the starting cell:  "))   
                got_start_index = True
            if not got_last_index:                                  #last excel cell to capture
                last_index = int(input("Enter the   last   cell:  "))        
                got_last_index = True
                getting_index = False
        except ValueError:
            print("\n   Oops! That wasn't a number. Please enter a number!\n ")
    
    while processing_names:                                         
        persons = start_program(start_index,last_index)             #beggining fuzzy search with given user indices
        processing_names = False

    start_again(start_index, last_index)                            #restarting search if user wants to


"""Kickstarging Program"""
if __name__ == "__main__":
    _main()






""" Notes for Improvements Later""""""

To increase flexibility of use, have user input the name of the file in the terminal.



"""










#These are deleted methods CAN ignore or delete later

"""This method will return a dictionary containing an abbridged version of the
    names stored in the Excel file """
"""def duplicate_check(first = [], mid = [], last = []):
    person = {}                         #dictionary to store account names
    #Checking for duplicates
    for i in range(37506):              #length of columns A,B,C
    #Start adding names to "person" dictionary
    name = first[i] + last[i]
    loop_mid = mid[i]               #middle name from for-loop
    if name not in person:          #create a dictionary of dictionaries w/ the following info
    person[name] = {
    "Middle"        : loop_mid,         #person's middle name
    "isDuplicate"   : False,            #indicates a duplicate match, default is set to False
    "pMatch"        : False,            #indicates a potential duplicate match
    "Duplicate Index": [],              #stores duplicate matches index
    "pMatch Index"  : [],               #stores potential duplicates index
    "Index"         : 'A' + str(i+2)    #index of the person
    }
    dict_mid = person[name]["Middle"]       #middle name from dictionary
    
    elif loop_mid == dict_mid or dict_mid[0] in loop_mid[0]:    #mid match completely or by first initial
    person[name]["isDuplicate"] = True                      #indicates duplicate of name
    person[name]["Duplicate Index"] += ['A' + str(i+2)]     #update match index of duplicate
    dict_mid  += loop_mid                                   #loop_mid added to dict_mid
    elif loop_mid == " ":                                   #loop middle name is empty
    person[name]["pMatch"] = True                       #potential match found
    person[name]["pMatch Index"] += ['A' + str(i+2)]    #update potential match index list
    return person
    """


"""
    This method will help read data that was pickled
        def unpickle(filename = ""):
        with open(filename, "rb") as f:
        while True:
        try:
        yield pickle.load(f)
        except EOFError:
        break
"""
